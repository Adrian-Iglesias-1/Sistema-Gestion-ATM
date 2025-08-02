import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime, time
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de la página
st.set_page_config(page_title="Sistema de Gestión ATM",
                   page_icon="🏧",
                   layout="wide")

# CSS mejorado con efectos modernos manteniendo el tema terminal
st.markdown("""
<style>
  /* Variables CSS para consistencia */
  :root {
    --primary-bg: #0d1117;
    --secondary-bg: #161b22;
    --accent-bg: #21262d;
    --primary-green: #39ff14;
    --secondary-green: #00d000;
    --dark-green: #004d00;
    --text-green: #7cfc00;
    --border-green: rgba(57, 255, 20, 0.3);
    --glow-green: 0 0 20px rgba(57, 255, 20, 0.3);
  }

  /* Fondo principal con gradiente sutil */
  .stApp {
    background: linear-gradient(135deg, var(--primary-bg) 0%, #0a0f1a 100%) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
  }

  /* Ocultar elementos de Streamlit */
  #MainMenu, footer, .stDeployButton {
    visibility: hidden !important;
  }

  /* Animaciones para títulos */
  @keyframes textGlow {
    from { text-shadow: 0 0 10px var(--primary-green), 0 0 20px var(--primary-green); }
    to { text-shadow: 0 0 20px var(--primary-green), 0 0 30px var(--primary-green), 0 0 40px var(--primary-green); }
  }

  @keyframes pulse {
    0% { transform: scale(1); }
    50% { transform: scale(1.02); }
    100% { transform: scale(1); }
  }

  /* Títulos principales */
  h1 {
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    color: var(--primary-green) !important;
    text-shadow: 
      0 0 10px var(--primary-green),
      0 0 20px var(--primary-green) !important;
    animation: textGlow 2s ease-in-out infinite alternate;
    font-weight: 700 !important;
    letter-spacing: 2px !important;
  }

  h2, h3 {
    color: var(--text-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    text-shadow: 0 0 5px var(--text-green) !important;
    letter-spacing: 1px !important;
  }

  /* Sidebar mejorado */
  .css-1lcbmhc {
    background: linear-gradient(180deg, var(--secondary-bg) 0%, var(--accent-bg) 100%) !important;
    border-right: 2px solid var(--primary-green) !important;
    box-shadow: inset -5px 0 15px rgba(57, 255, 20, 0.1) !important;
    padding: 1.5rem !important;
  }

  .css-1lcbmhc h1, .css-1lcbmhc h2, .css-1lcbmhc h3 {
    color: var(--primary-green) !important;
    font-size: 1.2rem !important;
    text-transform: uppercase !important;
    margin-bottom: 1rem !important;
  }

  /* Botones con efectos hover mejorados */
  .stButton > button {
    background: linear-gradient(135deg, var(--dark-green), var(--secondary-green)) !important;
    color: #ffffff !important;
    border: 2px solid var(--primary-green) !important;
    border-radius: 8px !important;
    padding: 0.75rem 1.5rem !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 1px !important;
    position: relative !important;
    overflow: hidden !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    cursor: pointer !important;
  }

  .stButton > button::before {
    content: '';
    position: absolute;
    top: 0;
    left: -100%;
    width: 100%;
    height: 100%;
    background: linear-gradient(90deg, transparent, rgba(255,255,255,0.2), transparent);
    transition: left 0.5s;
  }

  .stButton > button:hover::before {
    left: 100%;
  }

  .stButton > button:hover {
    background: linear-gradient(135deg, var(--secondary-green), var(--primary-green)) !important;
    box-shadow: var(--glow-green) !important;
    transform: translateY(-3px) scale(1.02) !important;
    border-color: var(--text-green) !important;
  }

  .stButton > button:active {
    transform: translateY(-1px) scale(0.98) !important;
  }

  /* Slider mejorado */
  .stSlider .css-1wy0on6 {
    background: linear-gradient(90deg, var(--dark-green), var(--secondary-green)) !important;
    border: 2px solid var(--border-green) !important;
    border-radius: 8px !important;
  }

  .stSlider .css-14xtw13 {
    color: var(--primary-green) !important;
    font-weight: 700 !important;
  }

  /* DataFrames con estilo matrix mejorado */
  .stDataFrame {
    border: 2px solid var(--border-green) !important;
    border-radius: 12px !important;
    overflow: hidden !important;
    box-shadow: var(--glow-green) !important;
    background: var(--primary-bg) !important;
  }

  .stDataFrame table {
    background-color: var(--primary-bg) !important;
    color: var(--text-green) !important;
    border-collapse: collapse !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-size: 0.85rem !important;
  }

  .stDataFrame th, .stDataFrame td {
    border: 1px solid var(--border-green) !important;
    padding: 12px 8px !important;
    text-align: center !important;
  }

  .stDataFrame thead th {
    background: linear-gradient(135deg, var(--dark-green), var(--secondary-green)) !important;
    color: #ffffff !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.5px !important;
    border-bottom: 3px solid var(--primary-green) !important;
    position: sticky !important;
    top: 0 !important;
    z-index: 10 !important;
  }

  .stDataFrame tbody tr:nth-child(even) {
    background-color: rgba(57, 255, 20, 0.03) !important;
  }

  .stDataFrame tbody tr:hover {
    background-color: rgba(57, 255, 20, 0.1) !important;
    transform: scale(1.001) !important;
    transition: all 0.2s ease !important;
    box-shadow: inset 0 0 10px rgba(57, 255, 20, 0.2) !important;
  }

  /* Selectbox mejorado */
  .stSelectbox div[data-baseweb="select"] {
    background: var(--secondary-bg) !important;
    border: 2px solid var(--border-green) !important;
    border-radius: 8px !important;
    color: var(--text-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    transition: all 0.3s ease !important;
  }

  .stSelectbox div[data-baseweb="select"]:hover {
    border-color: var(--primary-green) !important;
    box-shadow: 0 0 10px var(--border-green) !important;
  }

  .stSelectbox div[data-baseweb="select"] > div {
    color: var(--text-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
  }

  /* File uploader mejorado */
  .stFileUploader {
    border: 2px dashed var(--border-green) !important;
    border-radius: 12px !important;
    background: var(--secondary-bg) !important;
    transition: all 0.3s ease !important;
    padding: 1rem !important;
  }

  .stFileUploader:hover {
    border-color: var(--primary-green) !important;
    background: var(--accent-bg) !important;
    box-shadow: inset 0 0 20px rgba(57, 255, 20, 0.1) !important;
  }

  .stFileUploader label {
    color: var(--text-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 600 !important;
  }

  /* Tabs mejorados */
  .stTabs [data-baseweb="tab-list"] {
    background: var(--secondary-bg) !important;
    border-radius: 12px !important;
    padding: 0.5rem !important;
    border: 1px solid var(--border-green) !important;
  }

  .stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 8px !important;
    color: var(--text-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 600 !important;
    padding: 0.75rem 1.5rem !important;
    transition: all 0.3s ease !important;
    margin: 0 0.25rem !important;
  }

  .stTabs [data-baseweb="tab"]:hover {
    background: rgba(57, 255, 20, 0.1) !important;
    color: var(--primary-green) !important;
    transform: translateY(-2px) !important;
  }

  .stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, var(--dark-green), var(--secondary-green)) !important;
    color: #ffffff !important;
    box-shadow: 0 0 15px var(--border-green) !important;
    border: 1px solid var(--primary-green) !important;
  }

  /* Métricas mejoradas */
  .css-1xarl3l {
    background: var(--secondary-bg) !important;
    border: 1px solid var(--border-green) !important;
    border-radius: 8px !important;
    padding: 1rem !important;
    box-shadow: 0 4px 15px rgba(57, 255, 20, 0.1) !important;
  }

  .css-1xarl3l [data-testid="metric-container"] {
    background: transparent !important;
  }

  .css-1xarl3l [data-testid="metric-container"] > div {
    color: var(--primary-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 700 !important;
  }

  /* Scrollbar mejorado */
  ::-webkit-scrollbar {
    width: 12px !important;
    height: 12px !important;
  }

  ::-webkit-scrollbar-track {
    background: var(--secondary-bg) !important;
    border-radius: 6px !important;
  }

  ::-webkit-scrollbar-thumb {
    background: linear-gradient(45deg, var(--dark-green), var(--primary-green)) !important;
    border-radius: 6px !important;
    border: 2px solid var(--secondary-bg) !important;
  }

  ::-webkit-scrollbar-thumb:hover {
    background: linear-gradient(45deg, var(--secondary-green), var(--text-green)) !important;
    box-shadow: 0 0 10px var(--border-green) !important;
  }

  /* Mensajes de estado mejorados */
  .stSuccess {
    background: linear-gradient(135deg, rgba(57, 255, 20, 0.15), rgba(0, 208, 0, 0.1)) !important;
    border: 2px solid var(--primary-green) !important;
    border-radius: 8px !important;
    color: var(--primary-green) !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 600 !important;
    box-shadow: 0 0 15px rgba(57, 255, 20, 0.2) !important;
  }

  .stError {
    background: linear-gradient(135deg, rgba(255, 57, 57, 0.15), rgba(208, 0, 0, 0.1)) !important;
    border: 2px solid #ff3939 !important;
    border-radius: 8px !important;
    color: #ff6b6b !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 600 !important;
    box-shadow: 0 0 15px rgba(255, 57, 57, 0.2) !important;
  }

  .stWarning {
    background: linear-gradient(135deg, rgba(255, 193, 7, 0.15), rgba(255, 152, 0, 0.1)) !important;
    border: 2px solid #ffc107 !important;
    border-radius: 8px !important;
    color: #ffeb3b !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 600 !important;
    box-shadow: 0 0 15px rgba(255, 193, 7, 0.2) !important;
  }

  .stInfo {
    background: linear-gradient(135deg, rgba(33, 150, 243, 0.15), rgba(3, 169, 244, 0.1)) !important;
    border: 2px solid #2196f3 !important;
    border-radius: 8px !important;
    color: #64b5f6 !important;
    font-family: 'Courier New', 'Monaco', 'Menlo', monospace !important;
    font-weight: 600 !important;
    box-shadow: 0 0 15px rgba(33, 150, 243, 0.2) !important;
  }

  /* Spinner personalizado */
  .stSpinner > div {
    border-color: var(--border-green) !important;
    border-top-color: var(--primary-green) !important;
    border-width: 3px !important;
  }

  /* Progress bar mejorado */
  .stProgress .css-pq5sxo {
    background-color: var(--secondary-bg) !important;
    border-radius: 10px !important;
    border: 1px solid var(--border-green) !important;
  }

  .stProgress .css-1hr6zp3 {
    background: linear-gradient(90deg, var(--dark-green), var(--primary-green)) !important;
    border-radius: 8px !important;
    box-shadow: 0 0 10px var(--border-green) !important;
  }

  /* Contenedores con efecto glass */
  .element-container {
    backdrop-filter: blur(10px) !important;
    background: rgba(33, 38, 45, 0.6) !important;
    border: 1px solid var(--border-green) !important;
    border-radius: 12px !important;
    margin: 0.5rem 0 !important;
  }

  /* Efecto de fondo sutil */
  .stApp::before {
    content: '';
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background-image: 
      radial-gradient(circle at 25% 25%, rgba(57, 255, 20, 0.02) 0%, transparent 50%),
      radial-gradient(circle at 75% 75%, rgba(57, 255, 20, 0.02) 0%, transparent 50%);
    pointer-events: none;
    z-index: -1;
  }

  /* Separadores */
  hr {
    border: none !important;
    height: 2px !important;
    background: linear-gradient(90deg, transparent, var(--primary-green), transparent) !important;
    margin: 2rem 0 !important;
  }

  /* Responsive mejoras */
  @media (max-width: 768px) {
    .css-1lcbmhc {
      border-right: none !important;
      border-bottom: 2px solid var(--primary-green) !important;
    }

    h1 {
      font-size: 1.8rem !important;
    }

    .stButton > button {
      padding: 0.6rem 1.2rem !important;
      font-size: 0.9rem !important;
    }

    .stDataFrame {
      font-size: 0.75rem !important;
    }
  }

  /* Animación para elementos que aparecen */
  @keyframes fadeInUp {
    from {
      opacity: 0;
      transform: translateY(30px);
    }
    to {
      opacity: 1;
      transform: translateY(0);
    }
  }

  .element-container {
    animation: fadeInUp 0.6s ease-out;
  }
</style>
""",
            unsafe_allow_html=True)

# Inicializar estados de sesión para mejor UX
if 'processing' not in st.session_state:
    st.session_state.processing = False
if 'last_processed' not in st.session_state:
    st.session_state.last_processed = None
if 'resultados' not in st.session_state:
    st.session_state.resultados = {}


# Utilidades (mantengo toda la lógica original intacta)
def normalizar_id(series):
    return series.astype(str)\
                 .str.extract(r"(\d+)\s*$", expand=False)\
                 .str.lstrip('0')


def combinar_fecha_hora(f_val, h_val):
    try:
        fecha = pd.to_datetime(f_val, errors='coerce').date()
        if pd.isna(fecha):
            return None
        if pd.isna(h_val):
            return datetime.combine(fecha, time.min)
        hora = h_val if isinstance(h_val, time) else pd.to_datetime(
            h_val, errors='coerce').time()
        return datetime.combine(fecha, hora)
    except:
        return None


def limpiar_th_downtime(df_raw):
    header_idx = -1
    for i, row in df_raw.head(5).iterrows():
        texto = ' '.join(row.astype(str)).upper()
        if 'TICKET KEY' in texto and 'START TIME' in texto:
            header_idx = i
            break
    if header_idx < 0:
        return pd.DataFrame()
    df = df_raw.iloc[header_idx:].reset_index(drop=True)
    df.columns = df.iloc[0].astype(str).str.strip()
    df = df.drop(0).reset_index(drop=True)
    df.columns = df.columns.str.strip()
    return df.dropna(how='all').reset_index(drop=True)


# Funciones de categorización (sin cambios)
def categoria_por_sbif(codigo):
    try:
        c = str(int(float(codigo))).strip()
    except:
        c = str(codigo).strip()
    return {
        '2': 'Exigidos por SBIF',
        '6': 'Exigidos por SBIF',
        '7': 'Exigidos por SBIF',
        '5': 'Remodelación',
        '3': 'Vandalismo'
    }.get(c, 'Comunicaciones')


def categoria_por_resumen_falla(falla):
    m = str(falla).lower()
    mapa = {
        'dispensador con falla': 'Dispenser No Paga FLMG',
        'impresora de recibos': 'Impresora Recibos FLMG',
        'bna con falla': 'BNA/SDM/Deposito FLMG',
        '4 gavetas': '4 Gavetas Indisponibles',
        'host down': 'Aplicacion Fuera de Servicio',
        'comunicación con falla': 'Comunicaciones',
        'lector de tarjeta con falla': 'Lector de Tarjeta FLMG',
        'impresora sin papel': 'Sin Papel Recibos',
        'modo supervisor': 'Supervisor',
        'cash out': 'Cash Out'
    }
    for k, v in mapa.items():
        if k in m:
            return v
    return 'Comunicaciones'


def categoria_por_falla_ncr(falla):
    m = str(falla).lower()
    mapa = {
        'falla de configuración': 'Falla de HW / Servicio Técnico',
        'hardware': 'Falla de HW / Servicio Técnico',
        'pantalla con fallas': 'Falla de HW / Servicio Técnico',
        'lector de tarjeta con falla': 'Lector de Tarjeta SLMG',
        'impresora con falla': 'Impresora de recibos SLMG',
        'dispensador con falla': 'Dispenser no paga SLMG',
        'bna con falla': 'BNA/SDM/Deposito SLMG'
    }
    for k, v in mapa.items():
        if k in m:
            return v
    return 'Falla de HW / Servicio Técnico'


# Funciones de procesamiento (sin cambios en la lógica)
def procesar_exclusiones_cmm(df_cmm, df_th, tol):
    atm_col = 'ATM'
    fini = next(c for c in df_cmm
                if 'FECHA' in c.upper() and 'INICIO' in c.upper())
    hini = next(c for c in df_cmm
                if 'HORA' in c.upper() and 'INICIO' in c.upper())
    ffin = next(
        c for c in df_cmm
        if 'FECHA' in c.upper() and any(k in c.upper()
                                        for k in ['TERMINO', 'CIERRE', 'FIN']))
    hfin = next(
        c for c in df_cmm
        if 'HORA' in c.upper() and any(k in c.upper()
                                       for k in ['TERMINO', 'CIERRE', 'FIN']))
    sbif = next(c for c in df_cmm
                if 'SBIF' in c.upper() or 'CODIGO' in c.upper())

    df_cmm['_ini'] = df_cmm.apply(
        lambda r: combinar_fecha_hora(r[fini], r[hini]), axis=1)
    df_cmm['_fin'] = df_cmm.apply(
        lambda r: combinar_fecha_hora(r[ffin], r[hfin]), axis=1)

    df_th['id_norm'] = normalizar_id(df_th['ID'])
    df_th['ini_th'] = pd.to_datetime(df_th['START TIME'], errors='coerce')
    df_th['fin_th'] = pd.to_datetime(df_th['END TIME'], errors='coerce')

    out = []
    for _, r in df_cmm.iterrows():
        atm, ini, fin = r[atm_col], r['_ini'], r['_fin']
        if pd.isna(ini): continue
        orig = categoria_por_sbif(r[sbif])
        norm = normalizar_id(pd.Series(str(atm))).iloc[0]
        sub = df_th[df_th['id_norm'] == norm]

        if sub.empty:
            out.append({
                'ATM': atm,
                'Status Orig': orig,
                'Estado': 'No Encontrado',
                'TK TH': 'N/A',
                'Ini Orig': ini,
                'Fin Orig': fin,
                'Ini TH': pd.NaT,
                'Fin TH': pd.NaT
            })
        else:
            sub['diff'] = (sub['ini_th'] - ini).abs().dt.total_seconds() / 60
            best = sub.loc[sub['diff'].idxmin()]
            est = 'Encontrado' if best['diff'] <= tol else 'Diferencia'
            out.append({
                'ATM': atm,
                'Status Orig': orig,
                'Estado': est,
                'TK TH': best['TICKET KEY'],
                'Ini Orig': ini,
                'Fin Orig': fin,
                'Ini TH': best['ini_th'],
                'Fin TH': best['fin_th']
            })
    return pd.DataFrame(out)


def procesar_base_fallas(df_base, df_th):
    df_base['id_norm'] = normalizar_id(df_base['ATM'])
    df_base['Status'] = df_base['RESUMEN FALLA'].apply(
        categoria_por_resumen_falla)

    df_th2 = df_th.copy()
    df_th2['id_norm'] = normalizar_id(df_th2['ID'])
    df_th2['sd'] = pd.to_datetime(df_th2['START TIME'], errors='coerce')
    idx = df_th2.groupby('id_norm')['sd'].idxmax()
    df_lat = df_th2.loc[idx]
    df_lat['Inicio TH'] = df_lat['sd']
    df_lat['Fin TH'] = pd.to_datetime(df_lat['END TIME'], errors='coerce')

    m = pd.merge(df_base,
                 df_lat[['id_norm', 'TICKET KEY', 'Inicio TH', 'Fin TH']],
                 on='id_norm',
                 how='left')
    m['Estado'] = np.where(m['TICKET KEY'].notna(), 'Encontrado en TH',
                           'No Encontrado')
    m['TK TH'] = m['TICKET KEY'].fillna('N/A')
    return m[['ATM', 'TK TH', 'Status', 'Estado', 'Inicio TH', 'Fin TH']]


def procesar_base_fallas_ncr(df_ncr, df_th, tol=30):
    df_ncr['inicio'] = df_ncr.apply(lambda r: combinar_fecha_hora(
        r.get('FECHA INICIAL'), r.get('HORA INICIAL')),
                                    axis=1)
    df_th['id_norm'] = normalizar_id(df_th['ID'])
    df_th['inicio_th'] = pd.to_datetime(df_th['START TIME'], errors='coerce')
    df_th['fin_th'] = pd.to_datetime(df_th['END TIME'], errors='coerce')
    df_th['REFERENCE'] = df_th['REFERENCE'].astype(str).str.strip()

    out = []
    for _, r in df_ncr.iterrows():
        atm, wo, falla, ini = r['ATM'], str(
            r['WO']).strip(), r['FALLA NCR'], r['inicio']
        cat = categoria_por_falla_ncr(falla)
        est, tk, i_th, f_th = 'No Encontrado', 'N/A', pd.NaT, pd.NaT

        if wo.lower() not in ['nan', '']:
            dfw = df_th[df_th['REFERENCE'] == wo]
            if not dfw.empty:
                m0 = dfw.iloc[0]
                est, tk, i_th, f_th = 'Encontrado por WO', m0['REFERENCE'], m0[
                    'inicio_th'], m0['fin_th']

        if est == 'No Encontrado' and pd.notna(ini):
            norm = normalizar_id(pd.Series(str(atm))).iloc[0]
            sub = df_th[df_th['id_norm'] == norm].copy()
            if not sub.empty:
                sub['diff'] = (sub['inicio_th'] -
                               ini).abs().dt.total_seconds() / 60
                filt = sub[sub['diff'] <= tol]
                match = filt[filt['CATEGORY'].str.contains(cat,
                                                           case=False,
                                                           na=False)]
                if not match.empty:
                    b = match.loc[match['diff'].idxmin()]
                    est, tk, i_th, f_th = 'Encontrado (ID+Tiempo+Falla)', b[
                        'REFERENCE'], b['inicio_th'], b['fin_th']
                elif not filt.empty:
                    b = filt.loc[filt['diff'].idxmin()]
                    est, tk, i_th, f_th = 'Encontrado (ID+Tiempo)', b[
                        'REFERENCE'], b['inicio_th'], b['fin_th']
                else:
                    b = sub.iloc[0]
                    est, tk, i_th, f_th = 'Encontrado (Solo ID)', b[
                        'REFERENCE'], b['inicio_th'], b['fin_th']

        out.append({
            'ATM': atm,
            'TK TH': tk,
            'Status (Categoría)': cat,
            'Inicio TH': i_th,
            'Fin TH': f_th,
            'Estado Búsqueda': est
        })
    return pd.DataFrame(out)


# Función para validar archivos
def validate_files(file_dat, file_th):
    """Valida que los archivos sean correctos"""
    errors = []
    if not file_dat:
        errors.append("❌ Archivo de datos ATM es requerido")
    if not file_th:
        errors.append("❌ Archivo TH Downtime es requerido")
    return errors


# Interfaz principal mejorada
def main():
    # Header principal con métricas
    st.title("🏧 Sistema de Gestión ATM")
    st.markdown(
        "**🖥️ Terminal de Procesamiento de Datos v2.0** | *Análisis Automatizado de Exclusiones y Fallas*"
    )

    # Sidebar mejorado con validación visual
    with st.sidebar:
        st.header("🔧 Panel de Control")

        # Sección de archivos con validación visual
        st.subheader("📂 Carga de Archivos")

        file_dat = st.file_uploader(
            "📊 Datos ATM (Excel)",
            type=['xlsx', 'xls'],
            help="Archivo Excel con datos de ATMs para procesamiento")

        if file_dat:
            st.success("✅ Archivo de datos cargado correctamente")
            # Cargar hojas disponibles
            excel = pd.ExcelFile(file_dat)
            hojas = excel.sheet_names
        else:
            st.warning("⚠️ Archivo de datos requerido")
            excel, hojas = None, []

        file_th = st.file_uploader(
            "📉 TH Downtime (Excel)",
            type=['xlsx', 'xls'],
            help="Archivo Excel con datos de tiempo de inactividad")

        if file_th:
            st.success("✅ Archivo TH cargado correctamente")
        else:
            st.warning("⚠️ Archivo TH requerido")

        # Información del sistema
        st.markdown("---")
        st.subheader("📊 Estado del Sistema")

        # Métricas de estado
        col1, col2 = st.columns(2)
        with col1:
            if file_dat and file_th:
                st.metric("Estado", "🟢 Listo", "Archivos OK")
            else:
                st.metric("Estado", "🟡 Esperando", "Faltan archivos")

        with col2:
            st.metric("Hojas", len(hojas) if hojas else 0, "Disponibles")

        st.info(
            f"🕒 Última actualización: {datetime.now().strftime('%H:%M:%S')}")

        # Progress bar cuando se procesa
        if st.session_state.processing:
            st.markdown("---")
            st.subheader("⚡ Procesando...")
            progress_placeholder = st.empty()
            with progress_placeholder:
                st.progress(0.5)
                st.info("🔄 Analizando datos...")

    # Área principal con tabs mejorados
    tab1, tab2, tab3 = st.tabs(["⚙️ Configuración", "📊 Resultados", "📋 Ayuda"])

    with tab1:
        st.subheader("⚙️ Configuración de Procesamiento")

        # Validación de archivos
        validation_errors = validate_files(file_dat, file_th)
        if validation_errors:
            for error in validation_errors:
                st.error(error)
            st.info(
                "💡 **Instrución:** Carga ambos archivos en el panel lateral para continuar."
            )
            return

        # Configuración en layout organizado
        col1, col2 = st.columns([1, 1])

        with col1:
            st.markdown("**📋 Tipos de Procesamiento**")
            excl = st.selectbox(
                "🔄 Exclusiones-CMM", ["No procesar"] + hojas,
                key='excl',
                help="Selecciona la hoja con datos de exclusiones CMM")
            base = st.selectbox(
                "⚡ Base Fallas", ["No procesar"] + hojas,
                key='base',
                help="Selecciona la hoja con datos de fallas generales")
            ncr = st.selectbox(
                "🛠️ Base Fallas NCR", ["No procesar"] + hojas,
                key='ncr',
                help="Selecciona la hoja con datos de fallas NCR")

        with col2:
            st.markdown("**⚙️ Parámetros de Búsqueda**")
            tol = st.slider(
                "⏱️ Tolerancia (minutos)",
                min_value=0,
                max_value=120,
                value=30,
                step=5,
                key='tol',
                help=
                "Tolerancia en minutos para la búsqueda de coincidencias temporales"
            )

            st.markdown("**📊 Resumen de Configuración**")
            procesamiento_count = sum([
                excl != "No procesar", base != "No procesar", ncr
                != "No procesar"
            ])

            if procesamiento_count > 0:
                st.success(
                    f"✅ {procesamiento_count} procesamiento(s) configurado(s)")
                if excl != "No procesar":
                    st.info("🔄 Exclusiones-CMM: Activado")
                if base != "No procesar":
                    st.info("⚡ Base Fallas: Activado")
                if ncr != "No procesar":
                    st.info("🛠️ Base Fallas NCR: Activado")
            else:
                st.warning("⚠️ Selecciona al menos un tipo de procesamiento")

        # Botón de procesamiento mejorado
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            process_button = st.button(
                "🚀 INICIAR PROCESAMIENTO",
                use_container_width=True,
                disabled=not (file_dat and file_th
                              and procesamiento_count > 0))

        # Lógica de procesamiento (mantengo toda la funcionalidad original)
        if process_button:
            st.session_state.processing = True

            with st.spinner("🔄 Procesando datos..."):
                # Mostrar progreso
                progress_bar = st.progress(0)
                status_text = st.empty()

                try:
                    # Cargar y limpiar datos TH
                    status_text.text('📂 Cargando archivo TH Downtime...')
                    progress_bar.progress(20)

                    df_th_raw = pd.read_excel(file_th, header=None)
                    df_th = limpiar_th_downtime(df_th_raw)

                    if df_th.empty:
                        st.error(
                            "❌ No se pudo procesar el archivo TH Downtime. Verifica el formato."
                        )
                        st.session_state.processing = False
                        return

                    progress_bar.progress(40)
                    status_text.text('⚙️ Procesando datos...')

                    resultados = {}
                    progress_step = 60 / max(procesamiento_count, 1)
                    current_progress = 40

                    # Procesar Exclusiones-CMM
                    if excl != "No procesar":
                        status_text.text('🔄 Procesando Exclusiones-CMM...')
                        try:
                            if excel is not None:
                                resultados[
                                    'Exclusiones-CMM'] = procesar_exclusiones_cmm(
                                        excel.parse(excl), df_th, tol)
                            else:
                                st.error("❌ No se pudo cargar el archivo Excel")
                            current_progress += progress_step
                            progress_bar.progress(int(current_progress))
                        except Exception as e:
                            st.error(
                                f"❌ Error procesando Exclusiones-CMM: {str(e)}"
                            )

                    # Procesar Base Fallas
                    if base != "No procesar":
                        status_text.text('⚡ Procesando Base Fallas...')
                        try:
                            if excel is not None:
                                resultados['Base Fallas'] = procesar_base_fallas(
                                    excel.parse(base), df_th)
                            else:
                                st.error("❌ No se pudo cargar el archivo Excel")
                            current_progress += progress_step
                            progress_bar.progress(int(current_progress))
                        except Exception as e:
                            st.error(
                                f"❌ Error procesando Base Fallas: {str(e)}")

                    # Procesar Base Fallas NCR
                    if ncr != "No procesar":
                        status_text.text('🛠️ Procesando Base Fallas NCR...')
                        try:
                            if excel is not None:
                                resultados[
                                    'Base Fallas NCR'] = procesar_base_fallas_ncr(
                                        excel.parse(ncr), df_th, tol)
                            else:
                                st.error("❌ No se pudo cargar el archivo Excel")
                            current_progress += progress_step
                            progress_bar.progress(int(current_progress))
                        except Exception as e:
                            st.error(
                                f"❌ Error procesando Base Fallas NCR: {str(e)}"
                            )

                    # Finalizar procesamiento
                    status_text.text('✅ Procesamiento completado!')
                    progress_bar.progress(100)

                    # Guardar resultados en sesión
                    st.session_state.resultados = resultados
                    st.session_state.last_processed = datetime.now()
                    st.session_state.processing = False

                    if resultados:
                        st.success(
                            f"✅ **Procesamiento completado exitosamente!** Se procesaron {len(resultados)} tipo(s) de datos."
                        )
                        st.info(
                            "💡 **Próximo paso:** Ve a la pestaña 'Resultados' para ver y descargar los datos procesados."
                        )
                    else:
                        st.warning(
                            "⚠️ No se generaron resultados. Verifica la configuración."
                        )

                except Exception as e:
                    st.error(f"❌ **Error durante el procesamiento:** {str(e)}")
                    st.session_state.processing = False
                finally:
                    # Limpiar indicadores de progreso
                    progress_bar.empty()
                    status_text.empty()

    with tab2:
        st.subheader("📊 Resultados del Procesamiento")

        if st.session_state.resultados:
            # Información del último procesamiento
            if st.session_state.last_processed:
                st.info(
                    f"🕒 **Último procesamiento:** {st.session_state.last_processed.strftime('%d/%m/%Y %H:%M:%S')}"
                )

            # Mostrar resultados en sub-tabs
            result_tabs = st.tabs(list(st.session_state.resultados.keys()))

            for tab, (name,
                      df_out) in zip(result_tabs,
                                     st.session_state.resultados.items()):
                with tab:
                    st.markdown(f"### 📈 {name}")

                    # Métricas del resultado
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Registros", len(df_out))
                    with col2:
                        if 'Estado' in df_out.columns:
                            encontrados = len(
                                df_out[df_out['Estado'].str.contains(
                                    'Encontrado', na=False)])
                            st.metric("Encontrados", encontrados)
                        elif 'Estado Búsqueda' in df_out.columns:
                            encontrados = len(
                                df_out[df_out['Estado Búsqueda'].str.contains(
                                    'Encontrado', na=False)])
                            st.metric("Encontrados", encontrados)
                        else:
                            st.metric("Procesados", len(df_out))
                    with col3:
                        if 'Estado' in df_out.columns:
                            no_encontrados = len(
                                df_out[df_out['Estado'] == 'No Encontrado'])
                            st.metric("No Encontrados", no_encontrados)
                        elif 'Estado Búsqueda' in df_out.columns:
                            no_encontrados = len(df_out[
                                df_out['Estado Búsqueda'] == 'No Encontrado'])
                            st.metric("No Encontrados", no_encontrados)
                        else:
                            st.metric("Columnas", len(df_out.columns))

                    # Mostrar DataFrame
                    st.dataframe(df_out, use_container_width=True, height=400)

            # Botón de descarga mejorado
            st.markdown("---")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                # Generar archivo Excel con formato
                try:
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        # HOJA DE PORTADA PROFESIONAL
                        portada = pd.DataFrame({
                            'Información del Reporte': [
                                '',  # Espacio para título principal
                                '',  # Espacio para subtítulo
                                '',  # Línea separadora
                                f'Fecha de Generación: {datetime.now().strftime("%d/%m/%Y")}',
                                f'Hora de Generación: {datetime.now().strftime("%H:%M:%S")}',
                                '',
                                'RESUMEN DEL PROCESAMIENTO:',
                                f'• Tolerancia de búsqueda: {tol} minutos',
                                f'• Tipos de datos procesados: {len(st.session_state.resultados if "resultados" in st.session_state.__dict__ else {})}',
                                '• Exclusiones-CMM: Análisis de códigos SBIF',
                                '• Base Fallas: Categorización por tipo de falla',
                                '• Base Fallas NCR: Análisis especializado NCR',
                                '',
                                'ESTRUCTURA DEL REPORTE:',
                                '• Datos Originales: Información base del sistema',
                                '• Resultados del Análisis: Coincidencias encontradas',
                                '• Estado de Búsqueda: Encontrado/No Encontrado/Diferencia',
                                '• Información TH: Tickets de Help Desk relacionados',
                                '',
                                'Generado por: Sistema de Gestión ATM v2.0',
                                'Empresa: Análisis Automatizado de Datos ATM'
                            ]
                        })
                        portada.to_excel(writer, sheet_name='📋 PORTADA', index=False)
                        ws_portada = writer.sheets['📋 PORTADA']
                        
                        # FORMATO DE PORTADA PROFESIONAL
                        
                        # Título principal
                        ws_portada.merge_cells('A1:A3')
                        title_cell = ws_portada['A1']
                        title_cell.value = 'SISTEMA DE GESTIÓN ATM\nREPORTE DE ANÁLISIS INTEGRAL\nDATOS vs DOWNTIME'
                        title_cell.font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
                        title_cell.fill = PatternFill(start_color='1F4E79', end_color='2E5A87', fill_type='solid')
                        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Ajustar altura de las filas del título
                        ws_portada.row_dimensions[1].height = 30
                        ws_portada.row_dimensions[2].height = 30
                        ws_portada.row_dimensions[3].height = 30
                        
                        # Formato para información del reporte
                        for row_idx in range(4, len(portada) + 4):
                            cell = ws_portada.cell(row=row_idx, column=1)
                            cell_value = str(cell.value) if cell.value else ''
                            
                            if cell_value.startswith('RESUMEN') or cell_value.startswith('ESTRUCTURA'):
                                # Títulos de sección
                                cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                                cell.fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                ws_portada.row_dimensions[row_idx].height = 20
                            elif cell_value.startswith('•'):
                                # Elementos de lista
                                cell.font = Font(name='Arial', size=10, color='2F4F4F')
                                cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
                            elif cell_value.startswith('Fecha') or cell_value.startswith('Hora') or cell_value.startswith('Generado'):
                                # Información importante
                                cell.font = Font(name='Arial', size=11, bold=True, color='1F4E79')
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            else:
                                # Texto normal
                                cell.font = Font(name='Arial', size=10, color='333333')
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Ancho de columna y bordes
                        ws_portada.column_dimensions['A'].width = 60
                        
                        # Borde alrededor de toda la portada
                        for row_idx in range(1, len(portada) + 4):
                            cell = ws_portada.cell(row=row_idx, column=1)
                            cell.border = Border(
                                left=Side(style='medium', color='1F4E79'),
                                right=Side(style='medium', color='1F4E79'),
                                top=Side(style='thin', color='1F4E79'),
                                bottom=Side(style='thin', color='1F4E79')
                            )

                        # Agregar debugging para identificar el problema
                        print("=== DEBUGGING EXCEL GENERATION ===")
                        print(
                            f"st.session_state.resultados keys: {list(st.session_state.resultados.keys()) if 'resultados' in st.session_state.__dict__ else 'No existe resultados'}"
                        )

                        # Verificar que existan los resultados
                        if 'resultados' not in st.session_state.__dict__ or not st.session_state.resultados:
                            print("ERROR: No hay resultados para procesar")
                            st.error(
                                "No hay resultados para generar el Excel. Ejecuta el procesamiento primero."
                            )
                        else:
                            print(
                                f"Total de hojas a crear: {len(st.session_state.resultados)}"
                            )

                            # Hojas de resultados con formato mejorado y diferenciado
                            for name, df_out in st.session_state.resultados.items():
                                print(f"\n--- Procesando hoja: {name} ---")
                                print(f"Forma del df_out: {df_out.shape}")

                                try:
                                    # Obtener datos originales según el tipo
                                    df_in = None
                                    if name == 'Exclusiones-CMM' and excel is not None:
                                        df_in = excel.parse(excl)
                                        print(f"Exclusiones-CMM shape original: {df_in.shape}")
                                    elif name == 'Base Fallas' and excel is not None:
                                        df_in = excel.parse(base)
                                        print(f"Base Fallas shape original: {df_in.shape}")
                                        # Limpiar columnas innecesarias si existen
                                        if df_in.shape[1] > 14:
                                            df_in = df_in.drop(df_in.columns[11:15], axis=1)
                                    elif name == 'Base Fallas NCR' and excel is not None:
                                        df_in = excel.parse(ncr)
                                        print(f"Base Fallas NCR shape original: {df_in.shape}")
                                        # Limpiar columnas innecesarias si existen
                                        if df_in.shape[1] > 17:
                                            df_in = df_in.drop(df_in.columns[14:18], axis=1)

                                    # Combinar datos originales con resultados
                                    if df_in is not None:
                                        df_comb = pd.concat([
                                            df_in.reset_index(drop=True),
                                            df_out.reset_index(drop=True)
                                        ], axis=1)
                                        print(f"Combinando: df_in {df_in.shape} + df_out {df_out.shape} = {df_comb.shape}")
                                    else:
                                        df_comb = df_out.copy()
                                        print("Usando solo resultados")

                                    # Escribir al Excel con espacio para títulos
                                    print(f"Escribiendo hoja '{name}' al Excel...")
                                    df_comb.to_excel(writer, sheet_name=name, index=False, startrow=3)

                                    ws = writer.sheets[name]
                                    print(f"Hoja '{name}' creada exitosamente")

                                    # FORMATO PROFESIONAL PARA CLIENTE
                                    
                                    # 1. TÍTULO PRINCIPAL (Fila 1)
                                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df_comb.shape[1])
                                    title_cell = ws['A1']
                                    title_cell.value = f"REPORTE DE ANÁLISIS ATM - {name.upper()}"
                                    title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
                                    title_cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                                    
                                    # 2. SUBTÍTULOS SECCIÓN (Fila 2)
                                    if df_in is not None:
                                        # Subtítulo datos originales
                                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=df_in.shape[1])
                                        orig_cell = ws['A2']
                                        orig_cell.value = "DATOS ORIGINALES"
                                        orig_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                                        orig_cell.fill = PatternFill(start_color='8B4513', end_color='A0522D', fill_type='solid')
                                        orig_cell.alignment = Alignment(horizontal='center', vertical='center')
                                        
                                        # Subtítulo resultados
                                        start_col_results = df_in.shape[1] + 1
                                        ws.merge_cells(start_row=2, start_column=start_col_results, end_row=2, end_column=df_comb.shape[1])
                                        result_cell = ws.cell(row=2, column=start_col_results)
                                        result_cell.value = "RESULTADOS DEL ANÁLISIS"
                                        result_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                                        result_cell.fill = PatternFill(start_color='0066CC', end_color='4A90E2', fill_type='solid')
                                        result_cell.alignment = Alignment(horizontal='center', vertical='center')
                                    else:
                                        # Solo resultados
                                        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=df_comb.shape[1])
                                        result_cell = ws['A2']
                                        result_cell.value = "RESULTADOS DEL ANÁLISIS"
                                        result_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                                        result_cell.fill = PatternFill(start_color='0066CC', end_color='4A90E2', fill_type='solid')
                                        result_cell.alignment = Alignment(horizontal='center', vertical='center')

                                    # 3. FORMATO ENCABEZADOS DE COLUMNAS (Fila 4)
                                    for col_idx in range(1, df_comb.shape[1] + 1):
                                        cell = ws.cell(row=4, column=col_idx)
                                        if df_in is not None and col_idx <= df_in.shape[1]:
                                            # Encabezados datos originales - tonos marrones
                                            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                                            cell.fill = PatternFill(start_color='A0522D', end_color='CD853F', fill_type='solid')
                                        else:
                                            # Encabezados resultados - tonos azules
                                            cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                                            cell.fill = PatternFill(start_color='4A90E2', end_color='87CEEB', fill_type='solid')
                                        
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                        cell.border = Border(
                                            left=Side(style='thin', color='000000'),
                                            right=Side(style='thin', color='000000'),
                                            top=Side(style='thin', color='000000'),
                                            bottom=Side(style='medium', color='000000')
                                        )

                                    # 4. FORMATO DATOS (Desde fila 5)
                                    for row_idx in range(5, len(df_comb) + 5):
                                        for col_idx in range(1, df_comb.shape[1] + 1):
                                            cell = ws.cell(row=row_idx, column=col_idx)
                                            
                                            if df_in is not None and col_idx <= df_in.shape[1]:
                                                # Datos originales - tonos beige/marrones suaves
                                                if row_idx % 2 == 0:
                                                    cell.fill = PatternFill(start_color='F5F5DC', end_color='F5F5DC', fill_type='solid')  # Beige
                                                else:
                                                    cell.fill = PatternFill(start_color='FAEBD7', end_color='FAEBD7', fill_type='solid')  # Beige más claro
                                                cell.font = Font(name='Arial', size=9, color='2F4F4F')
                                            else:
                                                # Resultados - tonos azules suaves
                                                if row_idx % 2 == 0:
                                                    cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  # Azul muy claro
                                                else:
                                                    cell.fill = PatternFill(start_color='CCE7FF', end_color='CCE7FF', fill_type='solid')  # Azul claro
                                                cell.font = Font(name='Arial', size=9, color='003366', bold=True)
                                            
                                            cell.alignment = Alignment(horizontal='center', vertical='center')
                                            cell.border = Border(
                                                left=Side(style='thin', color='CCCCCC'),
                                                right=Side(style='thin', color='CCCCCC'),
                                                top=Side(style='thin', color='CCCCCC'),
                                                bottom=Side(style='thin', color='CCCCCC')
                                            )

                                    # 5. LÍNEA DIVISORIA VERTICAL entre originales y resultados
                                    if df_in is not None:
                                        separator_col = df_in.shape[1] + 1
                                        for row_idx in range(1, len(df_comb) + 5):
                                            cell = ws.cell(row=row_idx, column=separator_col)
                                            current_border = cell.border
                                            cell.border = Border(
                                                left=Side(style='medium', color='FF6600'),  # Línea naranja divisoria
                                                right=current_border.right,
                                                top=current_border.top,
                                                bottom=current_border.bottom
                                            )

                                    # 6. AJUSTAR ANCHO DE COLUMNAS automáticamente
                                    for col_idx in range(1, df_comb.shape[1] + 1):
                                        column_letter = get_column_letter(col_idx)
                                        max_length = 0
                                        
                                        # Revisar encabezado
                                        header_cell = ws.cell(row=4, column=col_idx)
                                        if header_cell.value:
                                            max_length = max(max_length, len(str(header_cell.value)))
                                        
                                        # Revisar datos (muestreo para performance)
                                        for row_idx in range(5, min(len(df_comb) + 5, 25)):
                                            cell_value = ws.cell(row=row_idx, column=col_idx).value
                                            if cell_value is not None:
                                                max_length = max(max_length, len(str(cell_value)))
                                        
                                        # Establecer ancho (mínimo 10, máximo 40)
                                        adjusted_width = min(max(max_length + 2, 10), 40)
                                        ws.column_dimensions[column_letter].width = adjusted_width

                                    # 7. CONGELAR PANELES para facilitar navegación
                                    ws.freeze_panes = 'A5'  # Congelar títulos y encabezados

                                    # 8. ALTURA DE FILAS para mejor presentación
                                    ws.row_dimensions[1].height = 25  # Título principal
                                    ws.row_dimensions[2].height = 20  # Subtítulos
                                    ws.row_dimensions[4].height = 18  # Encabezados

                                    print(f"Formato profesional aplicado a la hoja '{name}'")

                                except Exception as e:
                                    print(f"ERROR procesando hoja '{name}': {str(e)}")
                                    print(f"Tipo de error: {type(e).__name__}")
                                    import traceback
                                    traceback.print_exc()

                    print("=== FIN DEBUGGING ===")
                    buffer.seek(0)
                    st.download_button(
                        label="📥 DESCARGAR RESULTADOS FORMATEADOS",
                        data=buffer,
                        file_name=
                        f"Resultados_ATM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime=
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
                except Exception as e:
                    st.error(f"❌ Error generando archivo Excel: {str(e)}")
                    print(f"Error en Excel generation: {e}")
        else:
            st.info("🔄 **No hay resultados disponibles.**")
            st.markdown("""
            Para generar resultados:
            1. Ve a la pestaña **'Configuración'**
            2. Carga los archivos requeridos
            3. Selecciona los tipos de procesamiento
            4. Haz clic en **'INICIAR PROCESAMIENTO'**
            """)

    with tab3:
        st.subheader("📋 Guía de Uso del Sistema")

        # Instrucciones paso a paso
        st.markdown("""
        ### 🚀 **Pasos para usar el sistema:**

        #### 1. 📁 **Carga de Archivos**
        - **Datos ATM (Excel)**: Archivo principal con datos de ATMs
        - **TH Downtime (Excel)**: Archivo con registros de tiempo de inactividad
        - Los archivos deben estar en formato Excel (.xlsx o .xls)

        #### 2. ⚙️ **Configuración**
        - Selecciona las hojas a procesar para cada tipo de análisis
        - Ajusta la **tolerancia** (en minutos) para búsquedas temporales
        - El sistema validará automáticamente la configuración

        #### 3. 🎯 **Tipos de Procesamiento Disponibles**
        """)

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("""
            **🔄 Exclusiones-CMM**
            - Análisis de exclusiones por código SBIF
            - Búsqueda por ID de ATM y tiempo
            - Categorización automática por códigos
            """)

        with col2:
            st.markdown("""
            **⚡ Base Fallas**
            - Procesamiento de fallas generales
            - Matching con últimos registros TH
            - Categorización por tipo de falla
            """)

        with col3:
            st.markdown("""
            **🛠️ Base Fallas NCR**
            - Análisis específico de fallas NCR
            - Búsqueda por WO, ID y tiempo
            - Múltiples criterios de matching
            """)

        st.markdown("""
        #### 4. 📊 **Resultados**
        - Visualización en tablas interactivas
        - Métricas de resumen por cada procesamiento
        - Descarga en formato Excel con formato profesional

        #### 5. 📥 **Descarga**
        - Archivo Excel con múltiples hojas
        - Formato profesional con colores y estilos
        - Hoja de portada con información del reporte

        ---

        ### ⚙️ **Configuración de Tolerancia**

        La **tolerancia** determina qué tan precisas deben ser las coincidencias temporales:
        - **0-15 min**: Búsqueda muy estricta
        - **15-30 min**: Recomendado para la mayoría de casos
        - **30-60 min**: Para datos con variaciones temporales
        - **60+ min**: Búsqueda amplia (puede generar falsos positivos)

        ### 🔍 **Estados de Búsqueda**

        - **✅ Encontrado**: Coincidencia exacta encontrada
        - **🔍 Encontrado por WO**: Matching por Work Order
        - **⏱️ Encontrado (ID+Tiempo)**: Matching por ID y proximidad temporal
        - **❌ No Encontrado**: Sin coincidencias en los datos TH

        ### 💡 **Consejos y Buenas Prácticas**

        1. **Calidad de Datos**: Asegúrate de que los archivos tengan el formato esperado
        2. **Tolerancia Adecuada**: Comienza con 30 minutos y ajusta según resultados
        3. **Verificación**: Revisa las métricas de 'Encontrados' vs 'No Encontrados'
        4. **Backup**: Mantén copias de seguridad de tus archivos originales

        ### 🆘 **Solución de Problemas**

        **Error al cargar archivo TH:**
        - Verifica que contenga las columnas 'TICKET KEY' y 'START TIME'
        - Asegúrate de que no hay filas completamente vacías al inicio

        **Pocos resultados encontrados:**
        - Aumenta la tolerancia temporal
        - Verifica que los IDs de ATM coincidan en formato
        - Revisa las fechas y horas en ambos archivos

        **Archivo muy lento:**
        - Los archivos grandes pueden tomar varios minutos
        - El progreso se muestra en tiempo real
        - No cierres la ventana durante el procesamiento
        """)


if __name__ == "__main__":
    main()
