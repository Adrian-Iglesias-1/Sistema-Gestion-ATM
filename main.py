import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime, time
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuraci√≥n de la p√°gina
st.set_page_config(page_title="Sistema de Gesti√≥n ATM",
                   page_icon="üèß",
                   layout="wide")

# CSS simplificado para compatibilidad con Streamlit Cloud
st.markdown("""
<style>
  .stApp {
    background-color: #0d1117;
    color: #39ff14;
    font-family: 'Courier New', monospace;
  h1, h2, h3 {
    color: #39ff14;
    font-family: 'Courier New', monospace;
  }

  .stButton > button {
    background-color: #004d00;
    color: white;
    border: 1px solid #39ff14;
    border-radius: 5px;
  }
</style>
""",
            unsafe_allow_html=True)

# Inicializar estados de sesi√≥n para mejor UX
if 'processing' not in st.session_state:
    st.session_state.processing = False
if 'last_processed' not in st.session_state:
    st.session_state.last_processed = None
if 'resultados' not in st.session_state:
    st.session_state.resultados = {}


# Utilidades (mantengo toda la l√≥gica original intacta)
def normalizar_id(series):
    return series.astype(str)\
                 .str.extract(r"(\d+)\s*$", expand=False)\
                 .str.lstrip('0')


def combinar_fecha_hora(f_val, h_val):
    try:
        fecha = pd.to_datetime(f_val, errors='coerce').date()
        if pd.isna(fecha):
            return None
        if pd.isna(h_val):
            return datetime.combine(fecha, time.min)
        hora = h_val if isinstance(h_val, time) else pd.to_datetime(
            h_val, errors='coerce').time()
        return datetime.combine(fecha, hora)
    except:
        return None


def limpiar_th_downtime(df_raw):
    header_idx = -1
    for i, row in df_raw.head(5).iterrows():
        texto = ' '.join(row.astype(str)).upper()
        if 'TICKET KEY' in texto and 'START TIME' in texto:
            header_idx = i
            break
    if header_idx < 0:
        return pd.DataFrame()
    df = df_raw.iloc[header_idx:].reset_index(drop=True)
    df.columns = df.iloc[0].astype(str).str.strip()
    df = df.drop(0).reset_index(drop=True)
    df.columns = df.columns.str.strip()
    return df.dropna(how='all').reset_index(drop=True)


# Funciones de categorizaci√≥n (sin cambios)
def categoria_por_sbif(codigo):
    try:
        c = str(int(float(codigo))).strip()
    except:
        c = str(codigo).strip()
    return {
        '2': 'Exigidos por SBIF',
        '6': 'Exigidos por SBIF',
        '7': 'Exigidos por SBIF',
        '5': 'Remodelaci√≥n',
        '3': 'Vandalismo'
    }.get(c, 'Comunicaciones')


def categoria_por_resumen_falla(falla):
    m = str(falla).lower()
    mapa = {
        'dispensador con falla': 'Dispenser No Paga FLMG',
        'impresora de recibos': 'Impresora Recibos FLMG',
        'bna con falla': 'BNA/SDM/Deposito FLMG',
        '4 gavetas': '4 Gavetas Indisponibles',
        'host down': 'Aplicacion Fuera de Servicio',
        'comunicaci√≥n con falla': 'Comunicaciones',
        'lector de tarjeta con falla': 'Lector de Tarjeta FLMG',
        'impresora sin papel': 'Sin Papel Recibos',
        'modo supervisor': 'Supervisor',
        'cash out': 'Cash Out'
    }
    for k, v in mapa.items():
        if k in m:
            return v
    return 'Comunicaciones'


def categoria_por_falla_ncr(falla):
    m = str(falla).lower()
    mapa = {
        'falla de configuraci√≥n': 'Falla de HW / Servicio T√©cnico',
        'hardware': 'Falla de HW / Servicio T√©cnico',
        'pantalla con fallas': 'Falla de HW / Servicio T√©cnico',
        'lector de tarjeta con falla': 'Lector de Tarjeta SLMG',
        'impresora con falla': 'Impresora de recibos SLMG',
        'dispensador con falla': 'Dispenser no paga SLMG',
        'bna con falla': 'BNA/SDM/Deposito SLMG'
    }
    for k, v in mapa.items():
        if k in m:
            return v
    return 'Falla de HW / Servicio T√©cnico'


# Funciones de procesamiento (sin cambios en la l√≥gica)
def procesar_exclusiones_cmm(df_cmm, df_th, tol):
    atm_col = 'ATM'
    fini = next(c for c in df_cmm
                if 'FECHA' in c.upper() and 'INICIO' in c.upper())
    hini = next(c for c in df_cmm
                if 'HORA' in c.upper() and 'INICIO' in c.upper())
    ffin = next(
        c for c in df_cmm
        if 'FECHA' in c.upper() and any(k in c.upper()
                                        for k in ['TERMINO', 'CIERRE', 'FIN']))
    hfin = next(
        c for c in df_cmm
        if 'HORA' in c.upper() and any(k in c.upper()
                                       for k in ['TERMINO', 'CIERRE', 'FIN']))
    sbif = next(c for c in df_cmm
                if 'SBIF' in c.upper() or 'CODIGO' in c.upper())

    df_cmm['_ini'] = df_cmm.apply(
        lambda r: combinar_fecha_hora(r[fini], r[hini]), axis=1)
    df_cmm['_fin'] = df_cmm.apply(
        lambda r: combinar_fecha_hora(r[ffin], r[hfin]), axis=1)

    df_th['id_norm'] = normalizar_id(df_th['ID'])
    df_th['ini_th'] = pd.to_datetime(df_th['START TIME'], errors='coerce')
    df_th['fin_th'] = pd.to_datetime(df_th['END TIME'], errors='coerce')

    out = []
    for _, r in df_cmm.iterrows():
        atm, ini, fin = r[atm_col], r['_ini'], r['_fin']
        if pd.isna(ini): continue
        orig = categoria_por_sbif(r[sbif])
        norm = normalizar_id(pd.Series(str(atm))).iloc[0]
        sub = df_th[df_th['id_norm'] == norm]

        if sub.empty:
            out.append({
                'ATM': atm,
                'Status Orig': orig,
                'Estado': 'No Encontrado',
                'TK TH': 'N/A',
                'Ini Orig': ini,
                'Fin Orig': fin,
                'Ini TH': pd.NaT,
                'Fin TH': pd.NaT
            })
        else:
            sub['diff'] = (sub['ini_th'] - ini).abs().dt.total_seconds() / 60
            best = sub.loc[sub['diff'].idxmin()]
            est = 'Encontrado' if best['diff'] <= tol else 'Diferencia'
            out.append({
                'ATM': atm,
                'Status Orig': orig,
                'Estado': est,
                'TK TH': best['TICKET KEY'],
                'Ini Orig': ini,
                'Fin Orig': fin,
                'Ini TH': best['ini_th'],
                'Fin TH': best['fin_th']
            })
    return pd.DataFrame(out)


def procesar_base_fallas(df_base, df_th):
    df_base['id_norm'] = normalizar_id(df_base['ATM'])
    df_base['Status'] = df_base['RESUMEN FALLA'].apply(
        categoria_por_resumen_falla)

    df_th2 = df_th.copy()
    df_th2['id_norm'] = normalizar_id(df_th2['ID'])
    df_th2['sd'] = pd.to_datetime(df_th2['START TIME'], errors='coerce')
    idx = df_th2.groupby('id_norm')['sd'].idxmax()
    df_lat = df_th2.loc[idx]
    df_lat['Inicio TH'] = df_lat['sd']
    df_lat['Fin TH'] = pd.to_datetime(df_lat['END TIME'], errors='coerce')

    m = pd.merge(df_base,
                 df_lat[['id_norm', 'TICKET KEY', 'Inicio TH', 'Fin TH']],
                 on='id_norm',
                 how='left')
    m['Estado'] = np.where(m['TICKET KEY'].notna(), 'Encontrado en TH',
                           'No Encontrado')
    m['TK TH'] = m['TICKET KEY'].fillna('N/A')
    return m[['ATM', 'TK TH', 'Status', 'Estado', 'Inicio TH', 'Fin TH']]


def procesar_base_fallas_ncr(df_ncr, df_th, tol=30):
    df_ncr['inicio'] = df_ncr.apply(lambda r: combinar_fecha_hora(
        r.get('FECHA INICIAL'), r.get('HORA INICIAL')),
                                    axis=1)
    df_th['id_norm'] = normalizar_id(df_th['ID'])
    df_th['inicio_th'] = pd.to_datetime(df_th['START TIME'], errors='coerce')
    df_th['fin_th'] = pd.to_datetime(df_th['END TIME'], errors='coerce')
    df_th['REFERENCE'] = df_th['REFERENCE'].astype(str).str.strip()

    out = []
    for _, r in df_ncr.iterrows():
        atm, wo, falla, ini = r['ATM'], str(
            r['WO']).strip(), r['FALLA NCR'], r['inicio']
        cat = categoria_por_falla_ncr(falla)
        est, tk, i_th, f_th = 'No Encontrado', 'N/A', pd.NaT, pd.NaT

        if wo.lower() not in ['nan', '']:
            dfw = df_th[df_th['REFERENCE'] == wo]
            if not dfw.empty:
                m0 = dfw.iloc[0]
                est, tk, i_th, f_th = 'Encontrado por WO', m0['REFERENCE'], m0[
                    'inicio_th'], m0['fin_th']

        if est == 'No Encontrado' and pd.notna(ini):
            norm = normalizar_id(pd.Series(str(atm))).iloc[0]
            sub = df_th[df_th['id_norm'] == norm].copy()
            if not sub.empty:
                sub['diff'] = (sub['inicio_th'] -
                               ini).abs().dt.total_seconds() / 60
                filt = sub[sub['diff'] <= tol]
                match = filt[filt['CATEGORY'].str.contains(cat,
                                                           case=False,
                                                           na=False)]
                if not match.empty:
                    b = match.loc[match['diff'].idxmin()]
                    est, tk, i_th, f_th = 'Encontrado (ID+Tiempo+Falla)', b[
                        'REFERENCE'], b['inicio_th'], b['fin_th']
                elif not filt.empty:
                    b = filt.loc[filt['diff'].idxmin()]
                    est, tk, i_th, f_th = 'Encontrado (ID+Tiempo)', b[
                        'REFERENCE'], b['inicio_th'], b['fin_th']
                else:
                    b = sub.iloc[0]
                    est, tk, i_th, f_th = 'Encontrado (Solo ID)', b[
                        'REFERENCE'], b['inicio_th'], b['fin_th']

        out.append({
            'ATM': atm,
            'TK TH': tk,
            'Status (Categor√≠a)': cat,
            'Inicio TH': i_th,
            'Fin TH': f_th,
            'Estado B√∫squeda': est
        })
    return pd.DataFrame(out)


# Funci√≥n para validar archivos
def validate_files(file_dat, file_th):
    """Valida que los archivos sean correctos"""
    errors = []
    if not file_dat:
        errors.append("‚ùå Archivo de datos ATM es requerido")
    if not file_th:
        errors.append("‚ùå Archivo TH Downtime es requerido")
    return errors


# Interfaz principal mejorada
def main():
    # Header principal con m√©tricas
    st.title("üèß Sistema de Gesti√≥n ATM")
    st.markdown(
        "**üñ•Ô∏è Terminal de Procesamiento de Datos v2.0** | *An√°lisis Automatizado de Exclusiones y Fallas*"
    )

    # Sidebar mejorado con validaci√≥n visual
    with st.sidebar:
        st.header("üîß Panel de Control")

        # Secci√≥n de archivos con validaci√≥n visual
        st.subheader("üìÇ Carga de Archivos")

        file_dat = st.file_uploader(
            "üìä Datos ATM (Excel)",
            type=['xlsx', 'xls'],
            help="Archivo Excel con datos de ATMs para procesamiento")

        if file_dat:
            st.success("‚úÖ Archivo de datos cargado correctamente")
            # Cargar hojas disponibles
            excel = pd.ExcelFile(file_dat)
            hojas = excel.sheet_names
        else:
            st.warning("‚ö†Ô∏è Archivo de datos requerido")
            excel, hojas = None, []

        file_th = st.file_uploader(
            "üìâ TH Downtime (Excel)",
            type=['xlsx', 'xls'],
            help="Archivo Excel con datos de tiempo de inactividad")

        if file_th:
            st.success("‚úÖ Archivo TH cargado correctamente")
        else:
            st.warning("‚ö†Ô∏è Archivo TH requerido")

        # Informaci√≥n del sistema
        st.markdown("---")
        st.subheader("üìä Estado del Sistema")

        # M√©tricas de estado
        col1, col2 = st.columns(2)
        with col1:
            if file_dat and file_th:
                st.metric("Estado", "üü¢ Listo", "Archivos OK")
            else:
                st.metric("Estado", "üü° Esperando", "Faltan archivos")

        with col2:
            st.metric("Hojas", len(hojas) if hojas else 0, "Disponibles")

        st.info(
            f"üïí √öltima actualizaci√≥n: {datetime.now().strftime('%H:%M:%S')}")

        # Progress bar cuando se procesa
        if st.session_state.processing:
            st.markdown("---")
            st.subheader("‚ö° Procesando...")
            progress_placeholder = st.empty()
            with progress_placeholder:
                st.progress(0.5)
                st.info("üîÑ Analizando datos...")

    # √Årea principal con tabs mejorados
    tab1, tab2, tab3 = st.tabs(["‚öôÔ∏è Configuraci√≥n", "üìä Resultados", "üìã Ayuda"])

    with tab1:
        st.subheader("‚öôÔ∏è Configuraci√≥n de Procesamiento")

        # Validaci√≥n de archivos
        validation_errors = validate_files(file_dat, file_th)
        if validation_errors:
            for error in validation_errors:
                st.error(error)
            st.info(
                "üí° **Instruci√≥n:** Carga ambos archivos en el panel lateral para continuar."
            )
            return

        # Configuraci√≥n en layout organizado
        col1, col2 = st.columns([1, 1])

        with col1:
            st.markdown("**üìã Tipos de Procesamiento**")
            excl = st.selectbox(
                "üîÑ Exclusiones-CMM", ["No procesar"] + hojas,
                key='excl',
                help="Selecciona la hoja con datos de exclusiones CMM")
            base = st.selectbox(
                "‚ö° Base Fallas", ["No procesar"] + hojas,
                key='base',
                help="Selecciona la hoja con datos de fallas generales")
            ncr = st.selectbox(
                "üõ†Ô∏è Base Fallas NCR", ["No procesar"] + hojas,
                key='ncr',
                help="Selecciona la hoja con datos de fallas NCR")

        with col2:
            st.markdown("**‚öôÔ∏è Par√°metros de B√∫squeda**")
            tol = st.slider(
                "‚è±Ô∏è Tolerancia (minutos)",
                min_value=0,
                max_value=120,
                value=30,
                step=5,
                key='tol',
                help=
                "Tolerancia en minutos para la b√∫squeda de coincidencias temporales"
            )

            st.markdown("**üìä Resumen de Configuraci√≥n**")
            procesamiento_count = sum([
                excl != "No procesar", base != "No procesar", ncr
                != "No procesar"
            ])

            if procesamiento_count > 0:
                st.success(
                    f"‚úÖ {procesamiento_count} procesamiento(s) configurado(s)")
                if excl != "No procesar":
                    st.info("üîÑ Exclusiones-CMM: Activado")
                if base != "No procesar":
                    st.info("‚ö° Base Fallas: Activado")
                if ncr != "No procesar":
                    st.info("üõ†Ô∏è Base Fallas NCR: Activado")
            else:
                st.warning("‚ö†Ô∏è Selecciona al menos un tipo de procesamiento")

        # Bot√≥n de procesamiento mejorado
        st.markdown("---")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            process_button = st.button(
                "üöÄ INICIAR PROCESAMIENTO",
                use_container_width=True,
                disabled=not (file_dat and file_th
                              and procesamiento_count > 0))

        # L√≥gica de procesamiento (mantengo toda la funcionalidad original)
        if process_button:
            st.session_state.processing = True

            with st.spinner("üîÑ Procesando datos..."):
                # Mostrar progreso
                progress_bar = st.progress(0)
                status_text = st.empty()

                try:
                    # Cargar y limpiar datos TH
                    status_text.text('üìÇ Cargando archivo TH Downtime...')
                    progress_bar.progress(20)

                    df_th_raw = pd.read_excel(file_th, header=None)
                    df_th = limpiar_th_downtime(df_th_raw)

                    if df_th.empty:
                        st.error(
                            "‚ùå No se pudo procesar el archivo TH Downtime. Verifica el formato."
                        )
                        st.session_state.processing = False
                        return

                    progress_bar.progress(40)
                    status_text.text('‚öôÔ∏è Procesando datos...')

                    resultados = {}
                    progress_step = 60 / max(procesamiento_count, 1)
                    current_progress = 40

                    # Procesar Exclusiones-CMM
                    if excl != "No procesar":
                        status_text.text('üîÑ Procesando Exclusiones-CMM...')
                        try:
                            resultados[
                                'Exclusiones-CMM'] = procesar_exclusiones_cmm(
                                    excel.parse(excl), df_th, tol)
                            current_progress += progress_step
                            progress_bar.progress(int(current_progress))
                        except Exception as e:
                            st.error(
                                f"‚ùå Error procesando Exclusiones-CMM: {str(e)}"
                            )

                    # Procesar Base Fallas
                    if base != "No procesar":
                        status_text.text('‚ö° Procesando Base Fallas...')
                        try:
                            resultados['Base Fallas'] = procesar_base_fallas(
                                excel.parse(base), df_th)
                            current_progress += progress_step
                            progress_bar.progress(int(current_progress))
                        except Exception as e:
                            st.error(
                                f"‚ùå Error procesando Base Fallas: {str(e)}")

                    # Procesar Base Fallas NCR
                    if ncr != "No procesar":
                        status_text.text('üõ†Ô∏è Procesando Base Fallas NCR...')
                        try:
                            resultados[
                                'Base Fallas NCR'] = procesar_base_fallas_ncr(
                                    excel.parse(ncr), df_th, tol)
                            current_progress += progress_step
                            progress_bar.progress(int(current_progress))
                        except Exception as e:
                            st.error(
                                f"‚ùå Error procesando Base Fallas NCR: {str(e)}"
                            )

                    # Finalizar procesamiento
                    status_text.text('‚úÖ Procesamiento completado!')
                    progress_bar.progress(100)

                    # Guardar resultados en sesi√≥n
                    st.session_state.resultados = resultados
                    st.session_state.last_processed = datetime.now()
                    st.session_state.processing = False

                    if resultados:
                        st.success(
                            f"‚úÖ **Procesamiento completado exitosamente!** Se procesaron {len(resultados)} tipo(s) de datos."
                        )
                        st.info(
                            "üí° **Pr√≥ximo paso:** Ve a la pesta√±a 'Resultados' para ver y descargar los datos procesados."
                        )
                    else:
                        st.warning(
                            "‚ö†Ô∏è No se generaron resultados. Verifica la configuraci√≥n."
                        )

                except Exception as e:
                    st.error(f"‚ùå **Error durante el procesamiento:** {str(e)}")
                    st.session_state.processing = False
                finally:
                    # Limpiar indicadores de progreso
                    progress_bar.empty()
                    status_text.empty()

    with tab2:
        st.subheader("üìä Resultados del Procesamiento")

        if st.session_state.resultados:
            # Informaci√≥n del √∫ltimo procesamiento
            if st.session_state.last_processed:
                st.info(
                    f"üïí **√öltimo procesamiento:** {st.session_state.last_processed.strftime('%d/%m/%Y %H:%M:%S')}"
                )

            # Mostrar resultados en sub-tabs
            result_tabs = st.tabs(list(st.session_state.resultados.keys()))

            for tab, (name,
                      df_out) in zip(result_tabs,
                                     st.session_state.resultados.items()):
                with tab:
                    st.markdown(f"### üìà {name}")

                    # M√©tricas del resultado
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Registros", len(df_out))
                    with col2:
                        if 'Estado' in df_out.columns:
                            encontrados = len(
                                df_out[df_out['Estado'].str.contains(
                                    'Encontrado', na=False)])
                            st.metric("Encontrados", encontrados)
                        elif 'Estado B√∫squeda' in df_out.columns:
                            encontrados = len(
                                df_out[df_out['Estado B√∫squeda'].str.contains(
                                    'Encontrado', na=False)])
                            st.metric("Encontrados", encontrados)
                        else:
                            st.metric("Procesados", len(df_out))
                    with col3:
                        if 'Estado' in df_out.columns:
                            no_encontrados = len(
                                df_out[df_out['Estado'] == 'No Encontrado'])
                            st.metric("No Encontrados", no_encontrados)
                        elif 'Estado B√∫squeda' in df_out.columns:
                            no_encontrados = len(df_out[
                                df_out['Estado B√∫squeda'] == 'No Encontrado'])
                            st.metric("No Encontrados", no_encontrados)
                        else:
                            st.metric("Columnas", len(df_out.columns))

                    # Mostrar DataFrame
                    st.dataframe(df_out, use_container_width=True, height=400)

            # Bot√≥n de descarga mejorado
            st.markdown("---")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                # Generar archivo Excel con formato
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    # Hoja de portada
                    portada = pd.DataFrame({
                        'Sistema de Gesti√≥n ATM': [
                            'Reporte Generado',
                            f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}',
                            'Descripci√≥n: Resultados del procesamiento de Exclusiones-CMM, Base Fallas y Base Fallas NCR',
                            'Generado por: Sistema Automatizado v2.0',
                            f'Tolerancia utilizada: {tol} minutos',
                            f'Archivo: Resultados_ATM_Formateado.xlsx'
                        ]
                    })
                    portada.to_excel(writer, sheet_name='Portada', index=False)
                    ws_portada = writer.sheets['Portada']
                    ws_portada['A1'].font = Font(name='Arial',
                                                 size=16,
                                                 bold=True,
                                                 color='00FF00')
                    ws_portada['A1'].fill = PatternFill('solid',
                                                        fgColor='004D40')
                    for row in ws_portada['A2:A6']:
                        for cell in row:
                            cell.font = Font(name='Arial',
                                             size=12,
                                             color='000000')
                    ws_portada.column_dimensions['A'].width = 80

                    # Hojas de resultados con formato mejorado y diferenciado
                    for name, df_out in st.session_state.resultados.items():
                        # Obtener datos originales seg√∫n el tipo
                        if name == 'Exclusiones-CMM':
                            df_in = excel.parse(excl)
                        elif name == 'Base Fallas':
                            df_in = excel.parse(base)
                        else:  # Base Fallas NCR
                            df_in = excel.parse(ncr)

                        # Combinar datos originales con resultados
                        df_comb = pd.concat([
                            df_in.reset_index(drop=True),
                            df_out.reset_index(drop=True)
                        ],
                                            axis=1)

                        # Escribir al Excel
                        df_comb.to_excel(writer,
                                         sheet_name=name,
                                         index=False,
                                         startrow=2)
                        ws = writer.sheets[name]

                        # T√çTULO PRINCIPAL
                        ws['A1'] = name
                        ws['A1'].font = Font(name='Arial',
                                             size=16,
                                             bold=True,
                                             color='FFFFFF')
                        ws['A1'].fill = PatternFill('solid', fgColor='1F4E79')
                        ws['A1'].alignment = Alignment(horizontal='center',
                                                       vertical='center')
                        ws.merge_cells(start_row=1,
                                       start_column=1,
                                       end_row=1,
                                       end_column=df_comb.shape[1])

                        # SUBT√çTULOS PARA DIFERENCIAR SECCIONES
                        # Subt√≠tulo para datos originales
                        ws.merge_cells(start_row=2,
                                       start_column=1,
                                       end_row=2,
                                       end_column=df_in.shape[1])
                        ws['A2'] = "DATOS ORIGINALES"
                        ws['A2'].font = Font(name='Arial',
                                             size=12,
                                             bold=True,
                                             color='FFFFFF')
                        ws['A2'].fill = PatternFill(
                            'solid',
                            fgColor='8B4513')  # Marr√≥n para originales
                        ws['A2'].alignment = Alignment(horizontal='center',
                                                       vertical='center')

                        # Subt√≠tulo para resultados del procesamiento
                        start_col_results = df_in.shape[1] + 1
                        ws.merge_cells(start_row=2,
                                       start_column=start_col_results,
                                       end_row=2,
                                       end_column=df_comb.shape[1])
                        cell_results = ws.cell(row=2, column=start_col_results)
                        cell_results.value = "RESULTADOS DEL PROCESAMIENTO"
                        cell_results.font = Font(name='Arial',
                                                 size=12,
                                                 bold=True,
                                                 color='FFFFFF')
                        cell_results.fill = PatternFill(
                            'solid', fgColor='0066CC')  # Azul para resultados
                        cell_results.alignment = Alignment(horizontal='center',
                                                           vertical='center')

                        # ENCABEZADOS DE COLUMNAS (fila 3)
                        for col_idx, cell in enumerate(ws[3], 1):
                            if col_idx <= df_in.shape[1]:
                                # Encabezados de datos originales
                                cell.font = Font(name='Arial',
                                                 size=10,
                                                 bold=True,
                                                 color='FFFFFF')
                                cell.fill = PatternFill(
                                    'solid',
                                    fgColor='A0522D')  # Marr√≥n m√°s claro
                                cell.alignment = Alignment(horizontal='center',
                                                           vertical='center',
                                                           wrap_text=True)
                            else:
                                # Encabezados de resultados
                                cell.font = Font(name='Arial',
                                                 size=10,
                                                 bold=True,
                                                 color='FFFFFF')
                                cell.fill = PatternFill(
                                    'solid',
                                    fgColor='4A90E2')  # Azul m√°s claro
                                cell.alignment = Alignment(horizontal='center',
                                                           vertical='center',
                                                           wrap_text=True)

                            # Borde para todos los encabezados
                            cell.border = Border(left=Side(style='thin',
                                                           color='000000'),
                                                 right=Side(style='thin',
                                                            color='000000'),
                                                 top=Side(style='thin',
                                                          color='000000'),
                                                 bottom=Side(style='medium',
                                                             color='000000'))

                        # DATOS (desde fila 4 en adelante)
                        for row_idx in range(4, len(df_comb) + 4):
                            for col_idx in range(1, df_comb.shape[1] + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)

                                # Aplicar formato seg√∫n si es dato original o resultado
                                if col_idx <= df_in.shape[1]:
                                    # Datos originales - tonos beige/marrones claros
                                    if row_idx % 2 == 0:
                                        cell.fill = PatternFill(
                                            'solid',
                                            fgColor='F5F5DC')  # Beige claro
                                    else:
                                        cell.fill = PatternFill(
                                            'solid', fgColor='FAEBD7'
                                        )  # Beige m√°s claro
                                    cell.font = Font(
                                        name='Arial', size=9,
                                        color='2F4F4F')  # Texto gris oscuro
                                else:
                                    # Resultados del procesamiento - tonos azules claros
                                    if row_idx % 2 == 0:
                                        cell.fill = PatternFill(
                                            'solid',
                                            fgColor='E6F3FF')  # Azul muy claro
                                    else:
                                        cell.fill = PatternFill(
                                            'solid',
                                            fgColor='CCE7FF')  # Azul claro
                                    cell.font = Font(
                                        name='Arial',
                                        size=9,
                                        color='003366',
                                        bold=True
                                    )  # Texto azul oscuro y negrita

                                # Alineaci√≥n y bordes
                                cell.alignment = Alignment(horizontal='center',
                                                           vertical='center')
                                cell.border = Border(
                                    left=Side(style='thin', color='CCCCCC'),
                                    right=Side(style='thin', color='CCCCCC'),
                                    top=Side(style='thin', color='CCCCCC'),
                                    bottom=Side(style='thin', color='CCCCCC'))

                        # AJUSTAR ANCHO DE COLUMNAS
                        for col_idx in range(1, df_comb.shape[1] + 1):
                            # Calcular ancho basado en contenido
                            max_length = 0
                            column_letter = get_column_letter(col_idx)

                            # Revisar encabezado
                            header_length = len(
                                str(
                                    ws.cell(row=3, column=col_idx).value
                                    or ''))
                            max_length = max(max_length, header_length)

                            # Revisar datos (solo primeras 100 filas para performance)
                            for row_idx in range(4, min(len(df_comb) + 4,
                                                        104)):
                                cell_value = ws.cell(row=row_idx,
                                                     column=col_idx).value
                                if cell_value is not None:
                                    max_length = max(max_length,
                                                     len(str(cell_value)))

                            # Establecer ancho (m√≠nimo 12, m√°ximo 50)
                            adjusted_width = min(max(max_length + 2, 12), 50)
                            ws.column_dimensions[
                                column_letter].width = adjusted_width

                        # CONGELAR PANELES
                        ws.freeze_panes = 'A4'  # Congelar desde la fila 4 (despu√©s de t√≠tulos y encabezados)

                        # L√çNEA DIVISORIA VERTICAL entre datos originales y resultados
                        for row_idx in range(2, len(df_comb) + 4):
                            separator_col = df_in.shape[1] + 1
                            cell = ws.cell(row=row_idx, column=separator_col)
                            cell.border = Border(
                                left=Side(
                                    style='medium',
                                    color='FF6600'),  # L√≠nea naranja gruesa
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom)

                buffer.seek(0)

                st.download_button(
                    label="üì• DESCARGAR RESULTADOS FORMATEADOS",
                    data=buffer,
                    file_name=
                    f"Resultados_ATM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime=
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        else:
            st.info("üîÑ **No hay resultados disponibles.**")
            st.markdown("""
            Para generar resultados:
            1. Ve a la pesta√±a **'Configuraci√≥n'**
            2. Carga los archivos requeridos
            3. Selecciona los tipos de procesamiento
            4. Haz clic en **'INICIAR PROCESAMIENTO'**
            """)

    with tab3:
        st.subheader("üìã Gu√≠a de Uso del Sistema")

        # Instrucciones paso a paso
        st.markdown("""
        ### üöÄ **Pasos para usar el sistema:**

        #### 1. üìÅ **Carga de Archivos**
        - **Datos ATM (Excel)**: Archivo principal con datos de ATMs
        - **TH Downtime (Excel)**: Archivo con registros de tiempo de inactividad
        - Los archivos deben estar en formato Excel (.xlsx o .xls)

        #### 2. ‚öôÔ∏è **Configuraci√≥n**
        - Selecciona las hojas a procesar para cada tipo de an√°lisis
        - Ajusta la **tolerancia** (en minutos) para b√∫squedas temporales
        - El sistema validar√° autom√°ticamente la configuraci√≥n

        #### 3. üéØ **Tipos de Procesamiento Disponibles**
        """)

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("""
            **üîÑ Exclusiones-CMM**
            - An√°lisis de exclusiones por c√≥digo SBIF
            - B√∫squeda por ID de ATM y tiempo
            - Categorizaci√≥n autom√°tica por c√≥digos
            """)

        with col2:
            st.markdown("""
            **‚ö° Base Fallas**
            - Procesamiento de fallas generales
            - Matching con √∫ltimos registros TH
            - Categorizaci√≥n por tipo de falla
            """)

        with col3:
            st.markdown("""
            **üõ†Ô∏è Base Fallas NCR**
            - An√°lisis espec√≠fico de fallas NCR
            - B√∫squeda por WO, ID y tiempo
            - M√∫ltiples criterios de matching
            """)

        st.markdown("""
        #### 4. üìä **Resultados**
        - Visualizaci√≥n en tablas interactivas
        - M√©tricas de resumen por cada procesamiento
        - Descarga en formato Excel con formato profesional

        #### 5. üì• **Descarga**
        - Archivo Excel con m√∫ltiples hojas
        - Formato profesional con colores y estilos
        - Hoja de portada con informaci√≥n del reporte

        ---

        ### ‚öôÔ∏è **Configuraci√≥n de Tolerancia**

        La **tolerancia** determina qu√© tan precisas deben ser las coincidencias temporales:
        - **0-15 min**: B√∫squeda muy estricta
        - **15-30 min**: Recomendado para la mayor√≠a de casos
        - **30-60 min**: Para datos con variaciones temporales
        - **60+ min**: B√∫squeda amplia (puede generar falsos positivos)

        ### üîç **Estados de B√∫squeda**

        - **‚úÖ Encontrado**: Coincidencia exacta encontrada
        - **üîç Encontrado por WO**: Matching por Work Order
        - **‚è±Ô∏è Encontrado (ID+Tiempo)**: Matching por ID y proximidad temporal
        - **‚ùå No Encontrado**: Sin coincidencias en los datos TH

        ### üí° **Consejos y Buenas Pr√°cticas**

        1. **Calidad de Datos**: Aseg√∫rate de que los archivos tengan el formato esperado
        2. **Tolerancia Adecuada**: Comienza con 30 minutos y ajusta seg√∫n resultados
        3. **Verificaci√≥n**: Revisa las m√©tricas de 'Encontrados' vs 'No Encontrados'
        4. **Backup**: Mant√©n copias de seguridad de tus archivos originales

        ### üÜò **Soluci√≥n de Problemas**

        **Error al cargar archivo TH:**
        - Verifica que contenga las columnas 'TICKET KEY' y 'START TIME'
        - Aseg√∫rate de que no hay filas completamente vac√≠as al inicio

        **Pocos resultados encontrados:**
        - Aumenta la tolerancia temporal
        - Verifica que los IDs de ATM coincidan en formato
        - Revisa las fechas y horas en ambos archivos

        **Archivo muy lento:**
        - Los archivos grandes pueden tomar varios minutos
        - El progreso se muestra en tiempo real
        - No cierres la ventana durante el procesamiento
        """)


if __name__ == "__main__":
    main()
